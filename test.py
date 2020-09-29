import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('testFile.xlsx')
sheet = wb["Sheet1"]
# cell = sheet.cell(1, 2)
# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
	cell = sheet.cell(row, 3)
	new_price = cell.value * .9
	new_price_cell = sheet.cell(row, 4)
	new_price_cell.value = new_price

sheet.cell(1, 4).value = "new_price"

values = Reference(sheet,
				   min_row=2,
				   max_row=sheet.max_row,
				   min_col=4,
				   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "a6")

wb.save("testFile2.xlsx")
